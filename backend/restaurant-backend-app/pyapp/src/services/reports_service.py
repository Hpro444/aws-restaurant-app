"""Service for admin reporting dashboard table data."""

from __future__ import annotations

import csv
import io
from datetime import UTC, datetime, timedelta
from uuid import uuid4

import boto3
from botocore.exceptions import ClientError
from commons.app_config import AppConfig
from commons.log_helper import logger
from commons.report_utils import parse_date, pct_delta, period_start_for
from dto.reports import (
    CreateReportsDownloadRequest,
    DownloadFormat,
    GetReportsRequest,
    ReportsResponse,
    ReportType,
)
from repositories.feedback_cuisine_repository import FeedbackCuisineRepository
from repositories.feedback_service_repository import FeedbackServiceRepository
from repositories.location_report_repository import LocationReportRepository
from repositories.location_repository import LocationRepository
from repositories.order_repository import OrderRepository
from repositories.reservation_repository import ReservationRepository
from repositories.slot_repository import SlotRepository
from repositories.table_repository import TableRepository
from repositories.waiter_report_repository import WaiterReportRepository
from repositories.waiter_repository import WaiterRepository

_STAFF_COLUMN_LABELS: dict[str, str] = {
    "location": "Location",
    "waiter": "Waiter",
    "waiterEmail": "Waiter Email",
    "reportPeriodStart": "Period Start",
    "reportPeriodEnd": "Period End",
    "waiterWorkingHours": "Working Hours",
    "waiterOrdersProcessed": "Orders Processed",
    "deltaWaiterOrdersProcessedPct": "Delta Orders %",
    "averageServiceFeedback": "Avg Service Rating",
    "minimumServiceFeedback": "Min Service Rating",
    "deltaAverageServiceFeedbackPct": "Delta Rating %",
}

_SALES_COLUMN_LABELS: dict[str, str] = {
    "location": "Location",
    "reportPeriodStart": "Period Start",
    "reportPeriodEnd": "Period End",
    "ordersProcessed": "Orders Processed",
    "deltaOrdersProcessedPct": "Delta Orders %",
    "averageCuisineFeedback": "Avg Cuisine Rating",
    "minimumCuisineFeedback": "Min Cuisine Rating",
    "deltaAverageCuisineFeedbackPct": "Delta Cuisine Rating %",
    "revenue": "Revenue",
    "deltaRevenuePct": "Delta Revenue %",
}


class ReportsService:
    """Builds filtered report-table rows for the reporting UI."""

    def __init__(
        self,
        settings: AppConfig | None = None,
        waiter_report_repo: WaiterReportRepository | None = None,
        location_report_repo: LocationReportRepository | None = None,
        feedback_service_repo: FeedbackServiceRepository | None = None,
        feedback_cuisine_repo: FeedbackCuisineRepository | None = None,
        reservation_repo: ReservationRepository | None = None,
        slot_repo: SlotRepository | None = None,
        order_repo: OrderRepository | None = None,
        table_repo: TableRepository | None = None,
        waiter_repo: WaiterRepository | None = None,
        location_repo: LocationRepository | None = None,
    ) -> None:
        """Initialise repositories, creating defaults when omitted."""
        cfg = settings or AppConfig()
        self._settings = cfg
        self._waiter_report_repo = waiter_report_repo or WaiterReportRepository(cfg)
        self._location_report_repo = location_report_repo or LocationReportRepository(
            cfg
        )
        self._feedback_service_repo = (
            feedback_service_repo or FeedbackServiceRepository()
        )
        self._feedback_cuisine_repo = (
            feedback_cuisine_repo or FeedbackCuisineRepository()
        )
        self._reservation_repo = reservation_repo or ReservationRepository(cfg)
        self._slot_repo = slot_repo or SlotRepository(cfg)
        self._order_repo = order_repo or OrderRepository(cfg)
        self._table_repo = table_repo or TableRepository(cfg)
        self._waiter_repo = waiter_repo or WaiterRepository(cfg)
        self._location_repo = location_repo or LocationRepository(cfg)
        self._s3 = boto3.client("s3", region_name=cfg.aws_region)
        self._resolved_bucket_name: str | None = None

    def get_reports(self, request: GetReportsRequest) -> ReportsResponse:
        """Return report rows for requested type, period range, and optional location."""
        period_start, period_end = request.resolve_period_range()
        duration_days = (period_end - period_start).days
        prev_period_end = period_start - timedelta(days=1)
        prev_period_start = prev_period_end - timedelta(days=duration_days)

        current_week_starts = self._iter_week_starts(period_start, period_end)
        previous_week_starts = self._iter_week_starts(
            prev_period_start, prev_period_end
        )

        if request.report_type == ReportType.STAFF_PERFORMANCE:
            rows = self._build_staff_rows(
                request,
                period_start,
                period_end,
                current_week_starts,
                prev_period_start,
                prev_period_end,
                previous_week_starts,
            )
        else:
            rows = self._build_sales_rows(
                request,
                period_start,
                period_end,
                current_week_starts,
                prev_period_start,
                prev_period_end,
                previous_week_starts,
            )

        return ReportsResponse(
            report_type=request.report_type,
            period_start=period_start.isoformat(),
            period_end=period_end.isoformat(),
            rows=rows,
        )

    def _build_staff_rows(
        self,
        request: GetReportsRequest,
        period_start,
        period_end,
        current_week_starts: list[str],
        prev_period_start,
        prev_period_end,
        previous_week_starts: list[str],
    ) -> list[dict]:
        current_reports = self._load_waiter_reports(
            current_week_starts, period_start, period_end
        )
        previous_reports = self._load_waiter_reports(
            previous_week_starts,
            prev_period_start,
            prev_period_end,
        )

        if request.location_id is not None:
            current_reports = [
                report
                for report in current_reports
                if report.location_id == request.location_id
            ]
            previous_reports = [
                report
                for report in previous_reports
                if report.location_id == request.location_id
            ]

        current_agg = self._aggregate_waiter_reports(current_reports)
        previous_agg = self._aggregate_waiter_reports(previous_reports)

        keys = sorted(
            current_agg.keys(),
            key=lambda key: (
                (current_agg[key]["location_name"] or "").lower(),
                (current_agg[key]["waiter_last_name"] or "").lower(),
                (current_agg[key]["waiter_first_name"] or "").lower(),
            ),
        )

        rows = []
        for key in keys:
            current = current_agg[key]
            previous = previous_agg.get(key)

            current_avg = self._safe_average(
                current["service_feedback_sum"],
                current["service_feedback_count"],
            )
            previous_avg = (
                self._safe_average(
                    previous["service_feedback_sum"],
                    previous["service_feedback_count"],
                )
                if previous
                else None
            )

            rows.append(
                {
                    "location": current["location_name"],
                    "waiter": (
                        f"{current['waiter_first_name']} {current['waiter_last_name']}"
                    ).strip(),
                    "waiterEmail": current["waiter_email"],
                    "reportPeriodStart": period_start.isoformat(),
                    "reportPeriodEnd": period_end.isoformat(),
                    "waiterWorkingHours": self._format_number(current["working_hours"]),
                    "waiterOrdersProcessed": current["orders_processed"],
                    "deltaWaiterOrdersProcessedPct": self._format_pct(
                        pct_delta(
                            current["orders_processed"],
                            previous["orders_processed"] if previous else None,
                        )
                    ),
                    "averageServiceFeedback": self._format_number(current_avg),
                    "minimumServiceFeedback": self._format_number(
                        current["min_service_feedback"]
                    ),
                    "deltaAverageServiceFeedbackPct": self._format_pct(
                        pct_delta(current_avg, previous_avg)
                    ),
                }
            )

        return rows

    def _build_sales_rows(
        self,
        request: GetReportsRequest,
        period_start,
        period_end,
        current_week_starts: list[str],
        prev_period_start,
        prev_period_end,
        previous_week_starts: list[str],
    ) -> list[dict]:
        current_reports = self._load_location_reports(
            current_week_starts, period_start, period_end
        )
        previous_reports = self._load_location_reports(
            previous_week_starts,
            prev_period_start,
            prev_period_end,
        )

        if request.location_id is not None:
            current_reports = [
                report
                for report in current_reports
                if report.location_id == request.location_id
            ]
            previous_reports = [
                report
                for report in previous_reports
                if report.location_id == request.location_id
            ]

        current_agg = self._aggregate_location_reports(current_reports)
        previous_agg = self._aggregate_location_reports(previous_reports)

        keys = sorted(
            current_agg.keys(),
            key=lambda key: (current_agg[key]["location_name"] or "").lower(),
        )

        rows = []
        for key in keys:
            current = current_agg[key]
            previous = previous_agg.get(key)

            current_avg = self._safe_average(
                current["cuisine_feedback_sum"],
                current["cuisine_feedback_count"],
            )
            previous_avg = (
                self._safe_average(
                    previous["cuisine_feedback_sum"],
                    previous["cuisine_feedback_count"],
                )
                if previous
                else None
            )

            rows.append(
                {
                    "location": current["location_name"],
                    "reportPeriodStart": period_start.isoformat(),
                    "reportPeriodEnd": period_end.isoformat(),
                    "ordersProcessed": current["orders_processed"],
                    "deltaOrdersProcessedPct": self._format_pct(
                        pct_delta(
                            current["orders_processed"],
                            previous["orders_processed"] if previous else None,
                        )
                    ),
                    "averageCuisineFeedback": self._format_number(current_avg),
                    "minimumCuisineFeedback": self._format_number(
                        current["min_cuisine_feedback"]
                    ),
                    "deltaAverageCuisineFeedbackPct": self._format_pct(
                        pct_delta(current_avg, previous_avg)
                    ),
                    "revenue": self._format_number(current["revenue"]),
                    "deltaRevenuePct": self._format_pct(
                        pct_delta(
                            current["revenue"],
                            previous["revenue"] if previous else None,
                        )
                    ),
                }
            )

        return rows

    @staticmethod
    def _format_pct(value: float | None) -> str:
        """Render percentage values with sign and trailing percent symbol."""
        if value is None:
            return ""
        formatted = f"{value:.2f}".rstrip("0").rstrip(".")
        sign = "+" if value > 0 else ""
        return f"{sign}{formatted}%"

    @staticmethod
    def _format_number(value: float | int | None) -> str | int:
        """Render numeric values compactly while keeping empty values blank."""
        if value is None:
            return ""
        if isinstance(value, int):
            return value
        return f"{value:.2f}".rstrip("0").rstrip(".")

    @staticmethod
    def _iter_week_starts(period_start, period_end) -> list[str]:
        """Return ISO-week Monday dates that intersect requested date range."""
        current = period_start_for(period_start)
        last = period_start_for(period_end)
        starts = []
        while current <= last:
            starts.append(current.isoformat())
            current += timedelta(days=7)
        return starts

    @staticmethod
    def _is_report_within_period(report, period_start, period_end) -> bool:
        """Return True when report row overlaps requested date range."""
        report_start = parse_date(report.report_period_start)
        report_end = parse_date(report.report_period_end)
        return report_end >= period_start and report_start <= period_end

    def _load_waiter_reports(self, week_starts, period_start, period_end):
        """Load waiter report rows for requested week starts limited to period overlap.

        First attempts to compute reports dynamically from current database state.
        Falls back to seeded pre-computed reports if dynamic calculation fails.
        """
        # Try to use dynamic reports first (so updated feedbacks are reflected)
        dynamic_reports = self._compute_dynamic_waiter_reports(period_start, period_end)
        if dynamic_reports:
            return dynamic_reports

        # Fall back to seeded reports if dynamic calculation failed
        reports = []
        for week_start in week_starts:
            reports.extend(self._waiter_report_repo.find_by_period_start(week_start))
        return [
            report
            for report in reports
            if self._is_report_within_period(report, period_start, period_end)
        ]

    def _load_location_reports(self, week_starts, period_start, period_end):
        """Load location report rows for requested week starts limited to period overlap.

        First attempts to compute reports dynamically from current database state.
        Falls back to seeded pre-computed reports if dynamic calculation fails.
        """
        # Try to use dynamic reports first (so updated feedbacks are reflected)
        dynamic_reports = self._compute_dynamic_location_reports(
            period_start, period_end
        )
        if dynamic_reports:
            return dynamic_reports

        # Fall back to seeded reports if dynamic calculation failed
        reports = []
        for week_start in week_starts:
            reports.extend(self._location_report_repo.find_by_period_start(week_start))
        return [
            report
            for report in reports
            if self._is_report_within_period(report, period_start, period_end)
        ]

    @staticmethod
    def _aggregate_waiter_reports(reports):
        """Aggregate waiter weekly rows into a single row per waiter/location."""
        aggregated = {}
        for report in reports:
            key = (report.location_id, report.waiter_id)
            if key not in aggregated:
                aggregated[key] = {
                    "location_name": report.location_name,
                    "waiter_first_name": report.waiter_first_name,
                    "waiter_last_name": report.waiter_last_name,
                    "waiter_email": report.waiter_email,
                    "working_hours": 0.0,
                    "orders_processed": 0,
                    "service_feedback_count": 0,
                    "service_feedback_sum": 0.0,
                    "min_service_feedback": None,
                }

            row = aggregated[key]
            row["working_hours"] += report.working_hours or 0.0
            row["orders_processed"] += report.orders_processed or 0
            row["service_feedback_count"] += report.service_feedback_count or 0
            row["service_feedback_sum"] += report.service_feedback_sum or 0.0

            report_min = report.min_service_feedback
            if report_min is not None:
                row_min = row["min_service_feedback"]
                row["min_service_feedback"] = (
                    report_min if row_min is None else min(row_min, report_min)
                )

        return aggregated

    @staticmethod
    def _aggregate_location_reports(reports):
        """Aggregate location weekly rows into a single row per location."""
        aggregated = {}
        for report in reports:
            key = report.location_id
            if key not in aggregated:
                aggregated[key] = {
                    "location_name": report.location_name,
                    "orders_processed": 0,
                    "cuisine_feedback_count": 0,
                    "cuisine_feedback_sum": 0.0,
                    "min_cuisine_feedback": None,
                    "revenue": 0.0,
                }

            row = aggregated[key]
            row["orders_processed"] += report.orders_processed or 0
            row["cuisine_feedback_count"] += report.cuisine_feedback_count or 0
            row["cuisine_feedback_sum"] += report.cuisine_feedback_sum or 0.0
            row["revenue"] += report.revenue or 0.0

            report_min = report.min_cuisine_feedback
            if report_min is not None:
                row_min = row["min_cuisine_feedback"]
                row["min_cuisine_feedback"] = (
                    report_min if row_min is None else min(row_min, report_min)
                )

        return aggregated

    def _compute_dynamic_waiter_reports(
        self, period_start: datetime, period_end: datetime
    ) -> list:
        """Dynamically compute waiter reports from current database state.

        Aggregates data whose date falls within [period_start, period_end]:
        * orders_processed — orders linked to FINISHED reservations in period
        * working_hours — slots in period × 1.75h each
        * service_feedback — feedback entries in period

        Returns a list of dicts compatible with _aggregate_waiter_reports format.
        """
        from enums.reservation_status import ReservationStatus

        period_start_date = (
            period_start.date() if isinstance(period_start, datetime) else period_start
        )
        period_end_date = (
            period_end.date() if isinstance(period_end, datetime) else period_end
        )

        try:
            # Get all reservations and filter to FINISHED in period
            all_reservations = self._reservation_repo.scan()
            period_finished_ids = {
                r.id
                for r in all_reservations
                if r.status == ReservationStatus.FINISHED
                and period_start_date
                <= (
                    r.created_at.date()
                    if isinstance(r.created_at, datetime)
                    else r.created_at
                )
                <= period_end_date
            }

            # Get all orders and filter to those with FINISHED reservations
            all_orders = self._order_repo.scan()
            orders_by_waiter: dict = {}
            for order in all_orders:
                if order.reservation_id in period_finished_ids:
                    orders_by_waiter.setdefault(order.waiter_id, []).append(order)

            # Get all service feedbacks and filter by period
            all_feedbacks = self._feedback_service_repo.scan()
            feedbacks_by_waiter: dict = {}
            for fb in all_feedbacks:
                fb_date = fb.date.date() if isinstance(fb.date, datetime) else fb.date
                if period_start_date <= fb_date <= period_end_date:
                    feedbacks_by_waiter.setdefault(fb.waiter_id, []).append(fb)

            # Get all slots and filter by period and waiter
            all_slots = self._slot_repo.scan()
            slots_by_waiter: dict = {}
            for slot in all_slots:
                if slot.waiter_id is None:
                    continue
                slot_date = (
                    slot.date.date() if isinstance(slot.date, datetime) else slot.date
                )
                if period_start_date <= slot_date <= period_end_date:
                    slots_by_waiter.setdefault(slot.waiter_id, []).append(slot)

            # Get all waiters and locations
            all_waiters = self._waiter_repo.scan()
            all_locations = self._location_repo.scan()
            locations_by_id = {loc.id: loc for loc in all_locations}

            # Build dynamic reports for each waiter
            reports: list = []
            for waiter in all_waiters:
                w_orders = orders_by_waiter.get(waiter.id, [])
                w_feedbacks = feedbacks_by_waiter.get(waiter.id, [])
                w_slots = slots_by_waiter.get(waiter.id, [])

                # Calculate aggregates
                orders_processed = len(w_orders)
                unique_starts = {slot.start_time for slot in w_slots}
                working_hours = len(unique_starts) * 1.75

                fb_count = len(w_feedbacks)
                fb_sum = float(sum(f.rate for f in w_feedbacks if f.rate is not None))
                avg_fb = round(fb_sum / fb_count, 2) if fb_count else None
                min_fb = min(
                    (f.rate for f in w_feedbacks if f.rate is not None), default=None
                )

                location = locations_by_id.get(waiter.location_id)
                location_name = location.name if location else ""

                # Create a dict-like object that mimics WaiterReport for aggregation
                report = type(
                    "DynamicWaiterReport",
                    (),
                    {
                        "waiter_id": waiter.id,
                        "location_id": waiter.location_id,
                        "location_name": location_name,
                        "waiter_first_name": waiter.fname,
                        "waiter_last_name": waiter.lname,
                        "waiter_email": waiter.email,
                        "report_period_start": period_start_date.isoformat(),
                        "report_period_end": period_end_date.isoformat(),
                        "working_hours": working_hours,
                        "orders_processed": orders_processed,
                        "service_feedback_count": fb_count,
                        "service_feedback_sum": fb_sum,
                        "avg_service_feedback": avg_fb,
                        "min_service_feedback": min_fb,
                    },
                )()

                reports.append(report)

            return reports

        except Exception as e:
            logger.warning(
                "Failed to compute dynamic waiter reports, falling back to seeded data",
                error=str(e),
            )
            return []

    def _compute_dynamic_location_reports(
        self, period_start: datetime, period_end: datetime
    ) -> list:
        """Dynamically compute location reports from current database state.

        Aggregates cuisine feedback by location for the given period.

        Returns a list of dicts compatible with _aggregate_location_reports format.
        """
        from enums.reservation_status import ReservationStatus

        period_start_date = (
            period_start.date() if isinstance(period_start, datetime) else period_start
        )
        period_end_date = (
            period_end.date() if isinstance(period_end, datetime) else period_end
        )

        try:
            # Get all reservations and filter to FINISHED in period
            all_reservations = self._reservation_repo.scan()
            period_finished_ids = {
                r.id
                for r in all_reservations
                if r.status == ReservationStatus.FINISHED
                and period_start_date
                <= (
                    r.created_at.date()
                    if isinstance(r.created_at, datetime)
                    else r.created_at
                )
                <= period_end_date
            }

            # Get all orders and filter by period
            all_orders = self._order_repo.scan()
            orders_by_location: dict = {}
            for order in all_orders:
                if order.reservation_id in period_finished_ids:
                    # Get table info to find location
                    table = (
                        self._table_repo.get(order.table_id)
                        if hasattr(order, "table_id")
                        else None
                    )
                    if table:
                        location_id = table.location_id
                        orders_by_location.setdefault(location_id, []).append(order)

            # Get all cuisine feedbacks and filter by period
            all_feedbacks = self._feedback_cuisine_repo.scan()
            feedbacks_by_location: dict = {}
            for fb in all_feedbacks:
                fb_date = fb.date.date() if isinstance(fb.date, datetime) else fb.date
                if period_start_date <= fb_date <= period_end_date:
                    feedbacks_by_location.setdefault(fb.location_id, []).append(fb)

            # Get all locations
            all_locations = self._location_repo.scan()

            # Build dynamic reports for each location
            reports: list = []
            for location in all_locations:
                l_orders = orders_by_location.get(location.id, [])
                l_feedbacks = feedbacks_by_location.get(location.id, [])

                # Calculate aggregates
                orders_processed = len(l_orders)

                fb_count = len(l_feedbacks)
                fb_sum = float(sum(f.rate for f in l_feedbacks if f.rate is not None))
                avg_fb = round(fb_sum / fb_count, 2) if fb_count else None
                min_fb = min(
                    (f.rate for f in l_feedbacks if f.rate is not None), default=None
                )

                # Revenue calculation (sum of order totals)
                revenue = sum(
                    o.total_amount
                    for o in l_orders
                    if hasattr(o, "total_amount") and o.total_amount
                )

                # Create a dict-like object that mimics LocationReport for aggregation
                report = type(
                    "DynamicLocationReport",
                    (),
                    {
                        "location_id": location.id,
                        "location_name": location.name,
                        "report_period_start": period_start_date.isoformat(),
                        "report_period_end": period_end_date.isoformat(),
                        "orders_processed": orders_processed,
                        "cuisine_feedback_count": fb_count,
                        "cuisine_feedback_sum": fb_sum,
                        "avg_cuisine_feedback": avg_fb,
                        "min_cuisine_feedback": min_fb,
                        "revenue": revenue,
                    },
                )()

                reports.append(report)

            return reports

        except Exception as e:
            logger.warning(
                "Failed to compute dynamic location reports, falling back to seeded data",
                error=str(e),
            )
            return []

    @staticmethod
    def _safe_average(total: float, count: int) -> float | None:
        """Return rounded average for sum/count pairs or None when count is zero."""
        if count <= 0:
            return None
        return round(total / count, 2)

    # ------------------------------------------------------------------
    # Report file export
    # ------------------------------------------------------------------

    def _resolve_bucket_name(self) -> str:
        """Return the actual S3 bucket name, resolving it by alias on first call.

        Deployed bucket names carry the syndicate resources prefix/suffix
        (e.g. ``tm3-restaurant-reports-bucket-dev1``) while the configured
        alias is the bare name, so the bucket is located by listing buckets
        and matching on the alias as a substring — the same pattern
        DynamoRepository uses for table names. The result is cached for the
        Lambda container lifetime; falls back to the raw alias when no
        matching bucket is found.

        Returns:
            The fully-qualified bucket name as deployed.

        """
        if self._resolved_bucket_name:
            return self._resolved_bucket_name

        alias = self._settings.reports_bucket
        logger.info("Resolving S3 bucket name", alias=alias)

        try:
            response = self._s3.list_buckets()
        except ClientError as exc:
            logger.error("list_buckets failed", error=str(exc))
            self._resolved_bucket_name = alias
            return alias

        for bucket in response.get("Buckets", []):
            name = bucket.get("Name", "")
            if alias in name:
                self._resolved_bucket_name = name
                logger.info("Resolved bucket name", alias=alias, bucket=name)
                return name

        logger.info("No bucket found, falling back to alias", alias=alias)
        self._resolved_bucket_name = alias
        return alias

    def export_report_from_payload(self, request: CreateReportsDownloadRequest) -> str:
        """Build a report file from pre-computed rows and return a presigned download URL.

        Args:
            request: Validated request with report metadata, rows, and desired format.

        Returns:
            A temporary presigned URL for downloading the generated file.

        """
        label_map = (
            _STAFF_COLUMN_LABELS
            if request.report_type == ReportType.STAFF_PERFORMANCE
            else _SALES_COLUMN_LABELS
        )
        source_keys = list(label_map.keys())
        header_labels = [label_map[k] for k in source_keys]
        table_rows = [
            [self._to_text(row.get(k, "")) for k in source_keys] for row in request.rows
        ]

        filename_base = (
            f"report_{request.report_type.value}"
            f"_{request.period_start}_{request.period_end}"
        )

        fmt = request.download_format
        if fmt == DownloadFormat.CSV:
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerow(header_labels)
            writer.writerows(table_rows)
            content: bytes = buf.getvalue().encode("utf-8")
            content_type = "text/csv"
            filename = f"{filename_base}.csv"

        elif fmt == DownloadFormat.EXCEL:
            from openpyxl import Workbook  # noqa: PLC0415

            wb = Workbook()
            ws = wb.active
            ws.title = "Report"
            ws.append(header_labels)
            for row in table_rows:
                ws.append(row)
            buf_bytes = io.BytesIO()
            wb.save(buf_bytes)
            content = buf_bytes.getvalue()
            content_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = f"{filename_base}.xlsx"

        else:  # PDF
            content = self._build_pdf(header_labels, table_rows)
            content_type = "application/pdf"
            filename = f"{filename_base}.pdf"

        ts = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
        object_key = f"generated-reports/{ts}-{uuid4()}-{filename}"
        bucket_name = self._resolve_bucket_name()

        self._s3.put_object(
            Bucket=bucket_name,
            Key=object_key,
            Body=content,
            ContentType=content_type,
            ContentDisposition=f'attachment; filename="{filename}"',
        )

        return self._s3.generate_presigned_url(
            "get_object",
            Params={
                "Bucket": bucket_name,
                "Key": object_key,
                "ResponseContentType": content_type,
                "ResponseContentDisposition": f'attachment; filename="{filename}"',
            },
            ExpiresIn=self._settings.reports_url_expiration_seconds,
        )

    @staticmethod
    def _build_pdf(headers: list[str], rows: list[list[str]]) -> bytes:
        """Render report rows as a PDF table using fpdf2."""
        from fpdf import FPDF  # noqa: PLC0415

        pdf = FPDF(orientation="L", unit="mm", format="A4")
        pdf.set_auto_page_break(auto=True, margin=10)
        pdf.add_page()

        usable_width = 277.0  # A4 landscape minus default margins
        row_h = 7

        def _fit_text(value: str, width: float) -> str:
            """Trim text so it always fits inside a single PDF cell."""
            if width <= 0:
                return ""
            if pdf.get_string_width(value) <= width:
                return value
            ellipsis = "..."
            if pdf.get_string_width(ellipsis) > width:
                return ""
            trimmed = value
            while trimmed and pdf.get_string_width(trimmed + ellipsis) > width:
                trimmed = trimmed[:-1]
            return (trimmed + ellipsis) if trimmed else ellipsis

        # Compute natural column widths from headers + row values.
        pdf.set_font("Helvetica", "", 7)
        padding = 2.0
        min_col_w = 14.0
        col_widths: list[float] = []
        for idx, header in enumerate(headers):
            max_w = pdf.get_string_width(str(header)) + padding
            for row in rows:
                if idx < len(row):
                    max_w = max(max_w, pdf.get_string_width(str(row[idx])) + padding)
            col_widths.append(max(max_w, min_col_w))

        total_w = sum(col_widths)
        if total_w > usable_width and total_w > 0:
            scale = usable_width / total_w
            col_widths = [w * scale for w in col_widths]

        pdf.set_font("Helvetica", "B", 8)
        for idx, col in enumerate(headers):
            cell_w = col_widths[idx]
            cell_text = _fit_text(str(col), cell_w - 1.0)
            pdf.cell(cell_w, row_h, cell_text, border=1, align="C")
        pdf.ln()

        pdf.set_font("Helvetica", "", 7)
        for row in rows:
            for idx, val in enumerate(row):
                cell_w = col_widths[idx]
                cell_text = _fit_text(str(val), cell_w - 1.0)
                pdf.cell(cell_w, row_h, cell_text, border=1)
            pdf.ln()

        return bytes(pdf.output())

    @staticmethod
    def _to_text(value: object) -> str:
        """Convert any cell value to a string, returning blank for None."""
        if value is None:
            return ""
        return str(value)
